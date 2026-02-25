import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# Create workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ===== SHEET 1: Accounts =====
ws_accounts = wb.create_sheet('Accounts', 0)
accounts_headers = [
    'Account_ID', 'Account_Name', 'MRR_Current', 'ARR_Current', 'Renewal_Date',
    'Total_Licenses', 'Active_Users', 'Product_Usage_Percentage', 'Open_Tickets',
    'Avg_Resolution_Time', 'Last_Contact_Date'
]
ws_accounts.append(accounts_headers)

# Sample accounts data
accounts_data = [
    ['ACC001', 'Empresa Tech SA', 5000, 60000, (datetime.now() + timedelta(days=60)).strftime('%Y-%m-%d'), 100, 85, 78.5, 2, 2.5, (datetime.now() - timedelta(days=5)).strftime('%Y-%m-%d')],
    ['ACC002', 'Digital Solutions Inc', 8500, 102000, (datetime.now() + timedelta(days=120)).strftime('%Y-%m-%d'), 200, 140, 92.0, 1, 1.8, (datetime.now() - timedelta(days=2)).strftime('%Y-%m-%d')],
    ['ACC003', 'Cloud Services Ltd', 3200, 38400, (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d'), 75, 45, 55.3, 5, 4.2, (datetime.now() - timedelta(days=15)).strftime('%Y-%m-%d')],
    ['ACC004', 'Innovation Labs', 6800, 81600, (datetime.now() + timedelta(days=90)).strftime('%Y-%m-%d'), 150, 120, 88.7, 3, 3.1, (datetime.now() - timedelta(days=8)).strftime('%Y-%m-%d')],
    ['ACC005', 'Global Ventures', 4500, 54000, (datetime.now() + timedelta(days=150)).strftime('%Y-%m-%d'), 120, 60, 45.2, 8, 5.5, (datetime.now() - timedelta(days=45)).strftime('%Y-%m-%d')],
]

for row in accounts_data:
    ws_accounts.append(row)

# Format MRR columns as currency
for row in range(2, len(accounts_data) + 2):
    ws_accounts[f'C{row}'].number_format = '$#,##0.00'
    ws_accounts[f'D{row}'].number_format = '$#,##0.00'

# ===== SHEET 2: Period_Data =====
ws_period = wb.create_sheet('Period_Data', 1)
period_headers = [
    'Account_ID', 'Period', 'MRR_Starting', 'Expansion_Revenue', 'Contraction_Revenue',
    'Churned_Revenue', 'Clients_Start_Period', 'Clients_Churned', 
    'Clients_Eligible_for_Renewal', 'Clients_Renewed'
]
ws_period.append(period_headers)

period_data = [
    ['ACC001', '2025-Q4', 4800, 500, 100, 0, 80, 0, 16, 15],
    ['ACC001', '2026-Q1', 5200, 300, 50, 0, 85, 0, 12, 11],
    ['ACC002', '2025-Q4', 8000, 800, 100, 0, 150, 2, 25, 24],
    ['ACC002', '2026-Q1', 8600, 600, 200, 0, 160, 1, 20, 19],
    ['ACC003', '2025-Q4', 3000, 100, 50, 100, 60, 5, 10, 7],
    ['ACC003', '2026-Q1', 3100, 50, 100, 200, 55, 8, 10, 5],
    ['ACC004', '2025-Q4', 6500, 400, 50, 0, 120, 0, 20, 19],
    ['ACC004', '2026-Q1', 6800, 500, 100, 0, 130, 1, 18, 17],
    ['ACC005', '2025-Q4', 4200, 200, 150, 250, 100, 15, 15, 8],
    ['ACC005', '2026-Q1', 4400, 100, 200, 300, 95, 18, 15, 5],
]

for row in period_data:
    ws_period.append(row)

# Format currency columns
for row in range(2, len(period_data) + 2):
    for col in ['C', 'D', 'E', 'F']:
        ws_period[f'{col}{row}'].number_format = '$#,##0.00'

# ===== SHEET 3: NPS_Data =====
ws_nps = wb.create_sheet('NPS_Data', 2)
nps_headers = ['Account_ID', 'Period', 'NPS_Response']
ws_nps.append(nps_headers)

nps_data = [
    ['ACC001', '2025-Q4', 9],
    ['ACC001', '2026-Q1', 8],
    ['ACC002', '2025-Q4', 10],
    ['ACC002', '2026-Q1', 9],
    ['ACC002', '2026-Q1', 9],
    ['ACC003', '2025-Q4', 5],
    ['ACC003', '2026-Q1', 6],
    ['ACC004', '2025-Q4', 8],
    ['ACC004', '2026-Q1', 8],
    ['ACC004', '2026-Q1', 9],
    ['ACC005', '2025-Q4', 4],
    ['ACC005', '2026-Q1', 3],
    ['ACC005', '2026-Q1', 5],
]

for row in nps_data:
    ws_nps.append(row)

# Adjust column widths
for ws in [ws_accounts, ws_period, ws_nps]:
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width

# Save file
wb.save('datos_ejemplo.xlsx')
print("✓ Archivo Excel de ejemplo creado: datos_ejemplo.xlsx")
